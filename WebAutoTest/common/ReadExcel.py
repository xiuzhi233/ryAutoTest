"""
 -*- coding: utf-8 -*-
 @Project  : ryvueAutoTest
 @File     : ReadExcel.py
 @Author   : xiuzhi233
 @Time     : 2024/3/16 17:03
 @Describe : 
 
 
"""
import openpyxl

from WebAutoTest.common.ReadIni import ReadIni
from WebAutoTest.common.read_json import read_json
from WebAutoTest.data_config.settings import *


class ReadExcel:
    def __init__(self):
        """
        获取excel文件的路径和工作表名称，再进行读取
        """
        self.read_ini = ReadIni()
        excel_path = self.read_ini.get_file_path('CASE_DATA_XLSX')

        # 加载工作簿workbook
        wb = openpyxl.load_workbook(excel_path)
        work_sheet = self.read_ini.get_table_name('SHEET')
        self.ws = wb[work_sheet]

        # 读取用例的json数据
        case_data_path = self.read_ini.get_file_path("CASE_DATA_JSON")
        self.case_data_dict = read_json(case_data_path)

        expect_data_path = self.read_ini.get_file_path("EXPECT_DATA_JSON")
        self.expect_data_dict = read_json(expect_data_path)

        sql_data_path = self.read_ini.get_file_path("SQL_DATA_JSON")
        self.sql_data_dict = read_json(sql_data_path)

    def __get_cell_value(self, column: str, row: int) -> str:
        """
        获取指定单元格数据
        :param column:
        :param row:
        :return:
        """
        try:
            cell_value = self.ws[column + str(row)].value
        except ValueError:
            raise ValueError("传入的列号和行号错误!")
        else:
            if cell_value is None:
                return None
            elif cell_value.strip():
                return cell_value.strip()

    def case_number(self, row):
        return self.__get_cell_value(CASE_NUMBER, row)

    def module_name(self, row):
        return self.__get_cell_value(MODULE_NAME, row)

    def feature_name(self, row):
        return self.__get_cell_value(FEATURE_NAME, row)

    def case_title(self, row):
        return self.__get_cell_value(CASE_TITLE, row)

    def case_level(self, row):
        return self.__get_cell_value(CASE_LEVEL, row)

    def case_data(self, row):
        case_data_key = self.__get_cell_value(CASE_DATA, row)

        try:
            if case_data_key:
                module_name = self.module_name(row)
                feature_name = self.feature_name(row)

                return self.case_data_dict[module_name][feature_name][case_data_key]

        except KeyError:
            raise KeyError("请查看excel中的数据是否准确")

    def expect_data(self, row):
        expect_data_key = self.__get_cell_value(EXPECT_DATA, row)

        try:
            if expect_data_key:
                module_name = self.module_name(row)
                feature_name = self.feature_name(row)

                return self.expect_data_dict[module_name][feature_name][expect_data_key]

        except KeyError:
            raise KeyError("请查看excel中的数据是否准确")

    def sql_type(self, row):
        return self.__get_cell_value(SQL_TYPE, row)

    def sql_sentence(self, row):
        sql_data_key = self.__get_cell_value(SQL_SENTENCE, row)

        try:
            if sql_data_key:
                module_name = self.module_name(row)
                feature_name = self.feature_name(row)

                return self.sql_data_dict[module_name][feature_name][sql_data_key]

        except KeyError:
            raise KeyError("请查看excel中的数据是否准确")

    def update_key(self, row):
        return self.__get_cell_value(UPDATE_KEY, row)

    def get_data(self):
        """将所有的用例数据放入二维列表中"""
        list_data = list()
        for row in range(2, self.ws.max_row + 1):
            # 用例编号
            case_number = self.case_number(row)
            # 模块名称
            module_name = self.module_name(row)
            # 功能名称
            feature_name = self.feature_name(row)
            # 用例标题
            case_title = self.case_title(row)
            # 用例等级
            case_level = self.case_level(row)
            # 请求数据
            case_data = self.case_data(row)
            # 期望数据
            case_expect = self.expect_data(row)
            # sql语句类型
            sql_type = self.sql_type(row)
            # sql语句
            sql_sentence = self.sql_sentence(row)
            # 更新的key
            update_key = self.update_key(row)

            # 所有数据
            all_data = [case_number, module_name, feature_name, case_title, case_level, case_data, case_expect,
                        sql_type, sql_sentence, update_key]
            # 有5个None，规避空行
            if all_data.count(None) <= 5:
                list_data.append(all_data)

        else:
            return list_data


if __name__ == '__main__':
    excel = ReadExcel()
    print(excel.get_data()[22])
