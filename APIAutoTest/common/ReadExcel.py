"""
 -*- coding: utf-8 -*-
 @Project  : ryvueAutoTest
 @File     : ReadExcel.py
 @Author   : xiuzhi233
 @Time     : 2024/2/19 11:37
 @Describe : 
 读取excel文件
 
"""
import openpyxl
from APIAutoTest.common.ReadIni import ReadIni
from APIAutoTest.common.ReadJson import read_json
from APIAutoTest.data_config.settings import *


class ReadExcel:
    def __init__(self):
        """
        获取除了ini文件以外所有文件的路径，加载excel，获取工作表，读取所有json文件
        """
        self.read_ini = ReadIni()
        # 数据文件的路径
        case_data_path = self.read_ini.get_file_path("CASE_DATA_JSON")
        expect_data_path = self.read_ini.get_file_path("EXPECT_DATA_JSON")
        sql_data_path = self.read_ini.get_file_path("SQL_DATA_JSON")

        # 获取的用例数据
        self.case_data_dict = read_json(case_data_path)
        self.expect_data_dict = read_json(expect_data_path)
        self.sql_data_dict = read_json(sql_data_path)

        # 使用ReadIni对象读取excel路径
        excel_path = self.read_ini.get_file_path("CASE_DATA_XLSX")
        # 加载工作簿workbook
        wb = openpyxl.load_workbook(excel_path)
        # 获取工作表的名称table name
        table_name = self.read_ini.get_table_name("SHEET")
        self.ws = wb[table_name]

    def __get_cell_value(self, column: str, row: int) -> str:
        """
        根据指定的列和行获取单元格数据
        :param column:列
        :param row:行
        :return:单元格数据
        """
        # 获取单元格数据：工作表["行号列号"].value
        try:
            cell_value = self.ws[column + str(row)].value
        except ValueError:
            raise ValueError("传入的行号或列号错误！")
        else:
            # 判断value是否为none，如果为空就返回none
            if cell_value is None:
                return None
            # 去掉value的前后空格，如果是空的字符串返回none；否则返回的是一个值，这个值为空的字符串
            # 去掉value的前后空格，如果不是空的字符串，返回去掉空格后的字符串
            elif cell_value.strip():  # 非空为true       函数或方法的默认返回值为None
                return cell_value.strip()

    def case_number(self, row):
        """
        根据行号获取用例编号
        :param row:
        :return:
        """
        return self.__get_cell_value(CASE_NUMBER, row)

    def module_name(self, row):
        """
        根据行号获取模块名称
        :param row:行号
        :return:模块名称
        """
        # 直接调用私有方法来获取行号名称
        return self.__get_cell_value(MODULE_NAME, row)

    def api_name(self, row):
        """
        根据行号获取接口名称
        :param row:行号
        :return:模块名称
        """
        # 直接调用私有方法来获取行号名称
        return self.__get_cell_value(API_NAME, row)

    def case_title(self, row):
        """
        根据行号获取用例标题
        :param row:行号
        :return:模块名称
        """
        # 直接调用私有方法来获取行号名称
        return self.__get_cell_value(CASE_TITLE, row)

    def case_level(self, row):
        """
        根据行号获取用例等级
        :param row:
        :return:
        """
        return self.__get_cell_value(CASE_LEVEL, row)

    def req_method(self, row):
        """
        获取其请求方法
        :param row:
        :return:
        """
        return self.__get_cell_value(REQUEST_METHOD, row)

    def case_url(self, row):
        """
        获取url
        :param row:行号
        :return:url
        """
        host = self.read_ini.get_host("RYVUE")
        path = self.__get_cell_value(REQUEST_PATH, row)

        if path is not None:
            return host + path
        else:
            return None

    def req_MIME(self, row):
        """
        获取请求的媒体类型
        :param row:
        :return:
        """
        return self.__get_cell_value(MIME, row)

    def case_data(self, row):
        """
        获取用例数据
        :param row:
        :return:
        """
        # 先获取用例数据的key
        case_data_key = self.__get_cell_value(CASE_DATA, row)
        # 判断用例的数据是否为None，则返回None
        if case_data_key is None:
            return None
        # 不为空，用key去json文件提取数据
        else:
            # 获取json文件的路径
            case_data_path = self.read_ini.get_file_path("CASE_DATA_JSON")
            # 获取模块名称接口名称
            # 调用ReadJson
            json_data = read_json(case_data_path)
            module_name = self.module_name(row)
            api_name = self.api_name(row)
            # 提取出数据
            return json_data[module_name][api_name][case_data_key]

    def expect_data(self, row):
        """
        获取期望数据
        :param row:
        :return:
        """
        # 先获取用例数据的key
        expect_data_key = self.__get_cell_value(EXPECT_DATA, row)
        # 判断用例的数据是否为None，则返回None

        if expect_data_key is None:
            return None
        # 不为空，用key去json文件提取数据
        else:
            # 获取json文件的路径
            expect_data_path = self.read_ini.get_file_path("EXPECT_DATA_JSON")
            # 获取模块名称接口名称
            # 调用ReadJson
            json_data = read_json(expect_data_path)
            module_name = self.module_name(row)
            api_name = self.api_name(row)
            # 提取出数据
            return json_data[module_name][api_name][expect_data_key]

    def sql_type(self, row):
        """
        根据行号获取sql语句的类型
        :param row:
        :return:
        """
        # 获取sql语句的类型
        sql_type = self.__get_cell_value(SQL_TYPE, row)

        # 判断sql语句类型的值不为None，如果不为None，将sql语句的类型转大写返回，如果为None，返回None
        if sql_type:
            return sql_type.upper()
        else:
            return None

    def sql_data(self, row):
        """
        根据行号，获取sql语句
        :param row:
        :return:
        """
        sql_data_key = self.__get_cell_value(SQL_SENTENCE, row)

        # 判断sql语句的key是否不为None，如果是就提取sql语句，如果为None，返回None
        if sql_data_key:
            # 获取模块名称
            module_name = self.module_name(row)
            # 获取接口名称
            api_name = self.api_name(row)
            # 返回sql语句
            return self.sql_data_dict[module_name][api_name][sql_data_key]

    def update_key(self, row):
        """
        根据行号，获取更新的key
        :param row:
        :return:
        """
        return self.__get_cell_value(UPDATE_KEY, row)

    def get_data(self):
        """
        将需要的数据存放到一个二维列表中
        :return:
        """
        # 创建一个空列表，用于存放所有的数据
        list_data = list()
        for row in range(2, self.ws.max_row + 1):
            # 用例编号
            case_number = self.case_number(row)
            # 模块名称
            module_name = self.module_name(row)
            # 接口名称
            api_name = self.api_name(row)
            # 用例标题
            case_title = self.case_title(row)
            # 用例等级
            case_level = self.case_level(row)
            # 请求路径
            request_url = self.case_url(row)
            # 请求方法
            request_method = self.req_method(row)
            # 请求媒体类型
            request_mime = self.req_MIME(row)
            # 请求数据
            request_data = self.case_data(row)
            # 期望数据
            request_expect = self.expect_data(row)
            # sql语句类型
            sql_type = self.sql_type(row)
            # sql语句
            sql_sentence = self.sql_data(row)
            # 更新的key
            update_key = self.update_key(row)

            # 将要用的数据全部放入row_list中
            row_list = [case_number, module_name, api_name, case_title, case_level, request_url, request_method,
                        request_mime, request_data, request_expect, sql_type, sql_sentence, update_key]
            # 有5个None，规避空行
            if row_list.count(None) <= 6:
                list_data.append(row_list)
        else:
            return list_data


if __name__ == '__main__':
    read_excel = ReadExcel()
    print(read_excel.get_data())
